from openpyxl import load_workbook
from io import BytesIO

from src.electrical_calc.spreadsheets import (
    create_input_template,
    create_project_export,
    parse_circuit_workbook,
)


def test_template_round_trip():
    content = create_input_template()
    workbook = load_workbook(BytesIO(content))
    assert workbook.sheetnames == ["回路导入", "字段说明"]
    rows, errors = parse_circuit_workbook(content)
    assert not errors
    assert rows[0]["code"] == "AL-01"


def test_valid_import_and_duplicate_error():
    content = create_input_template()
    workbook = load_workbook(BytesIO(content))
    sheet = workbook["回路导入"]
    sheet["K5"] = 40
    sheet["L5"] = 4.61
    sheet["M5"] = 0.08
    output = BytesIO()
    workbook.save(output)
    rows, errors = parse_circuit_workbook(output.getvalue())
    assert not errors
    assert rows[0]["code"] == "AL-01"
    assert rows[0]["voltage_v"] == 220

    workbook = load_workbook(BytesIO(output.getvalue()))
    sheet = workbook["回路导入"]
    for cell in sheet[5]:
        sheet.cell(6, cell.column).value = cell.value
    duplicate = BytesIO()
    workbook.save(duplicate)
    rows, errors = parse_circuit_workbook(duplicate.getvalue())
    assert any("重复" in item for item in errors)


def test_project_export_has_expected_sheets():
    project = {"code": "P-01", "name": "测试", "description": ""}
    content = create_project_export(project, [], [], [])
    workbook = load_workbook(BytesIO(content), data_only=False)
    assert workbook.sheetnames == ["项目汇总", "回路输入快照", "计算记录", "结果明细", "计算过程", "警告清单", "计算输入快照", "依据清单"]
    assert workbook["项目汇总"]["B5"].value == 0
