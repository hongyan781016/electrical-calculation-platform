from io import BytesIO

from fastapi.testclient import TestClient
from openpyxl import load_workbook
import pytest

from src.electrical_calc.database import Database
from src.electrical_calc import web


def capture_template_context(monkeypatch):
    captured = {}
    original = web.templates.TemplateResponse

    def capture(*args, **kwargs):
        context = kwargs.get("context")
        if context is None and len(args) >= 3:
            context = args[2]
        captured.update(context or {})
        return original(*args, **kwargs)

    monkeypatch.setattr(web.templates, "TemplateResponse", capture)
    return captured


def test_health_and_project_flow(tmp_path, monkeypatch):
    monkeypatch.setattr(web, "db", Database(tmp_path / "web.db"))
    client = TestClient(web.app)
    assert client.get("/health").json() == {"status": "ok", "version": "0.4.0"}

    response = client.post(
        "/projects",
        data={"code": "WEB-01", "name": "网页测试项目", "description": ""},
        follow_redirects=False,
    )
    assert response.status_code == 303
    project_url = response.headers["location"]
    page = client.get(project_url)
    assert page.status_code == 200
    assert "网页测试项目" in page.text

    response = client.post(
        project_url + "/circuits",
        data={
            "code": "AL-01",
            "name": "照明",
            "phase": "1",
            "voltage_v": "220",
            "installed_power_kw": "5",
            "demand_factor": "0.8",
            "power_factor": "0.9",
            "efficiency": "1",
        },
        follow_redirects=False,
    )
    assert response.status_code == 303
    circuit = web.db.list_circuits(1)[0]
    response = client.post(f"/circuits/{circuit['id']}/calculate", follow_redirects=False)
    assert response.status_code == 303
    run_page = client.get(response.headers["location"])
    assert run_page.status_code == 200
    assert "负荷与选型" in run_page.text


def test_rule_cannot_be_approved_without_source(tmp_path, monkeypatch):
    monkeypatch.setattr(web, "db", Database(tmp_path / "web.db"))
    client = TestClient(web.app)
    rule = web.db.list_rules()[0]
    original_status = rule["status"]
    response = client.post(
        f"/rules/{rule['id']}",
        data={"name": rule["name"], "status": "approved"},
        follow_redirects=False,
    )
    assert response.status_code == 303
    assert web.db.list_rules()[0]["status"] == original_status


def test_quick_selection_does_not_require_project_or_excel(tmp_path, monkeypatch):
    monkeypatch.setattr(web, "db", Database(tmp_path / "quick.db"))
    client = TestClient(web.app)

    before_rules = len(web.db.list_rules())
    page = client.get("/quick")
    assert page.status_code == 200
    assert "0.4kV 回路计算" in page.text
    assert 'value="380"' in page.text
    assert 'name="circuit_role"' in page.text
    assert 'name="conductor_basis"' not in page.text
    assert 'name="cable_r_ohm_per_km"' not in page.text
    assert 'name="cable_x_ohm_per_km"' not in page.text
    assert 'name="power_factor"' in page.text
    assert 'name="prospective_short_circuit_ka"' not in page.text
    assert "上级变压器资料" in page.text
    assert 'name="short_circuit_method"' in page.text
    assert 'name="transformer_capacity_kva"' in page.text
    assert 'name="transformer_uk_percent"' in page.text
    assert 'name="fault_transformer_series_code"' in page.text
    assert 'name="fault_transformer_capacity_kva"' in page.text
    assert 'name="fault_fourth_conductor_role"' in page.text
    assert 'name="fault_busway_series_code"' in page.text
    assert 'name="fault_busway_rating_a"' in page.text
    assert 'name="pe_thermal_enabled"' not in page.text
    assert 'name="phase_thermal_enabled"' not in page.text
    assert "热稳定约束由系统自动形成" in page.text
    assert 'name="fault_clearing_time_s"' in page.text
    assert 'name="voltage_drop_limit_pct"' not in page.text
    assert "允许电压降：系统自动采用" in page.text
    assert 'name="installation_temperature_c"' in page.text
    assert 'name="enclosed_grouping_circuit_count"' in page.text
    assert "空气中敷设采用表6.22；埋地管槽采用表6.24" in page.text
    assert 'name="soil_thermal_resistivity_k_m_per_w"' in page.text
    assert 'name="buried_circuit_count"' in page.text
    assert 'name="buried_duct_spacing_m"' in page.text
    assert 'name="buried_depth_m"' in page.text
    assert ">埋地管槽<" in page.text
    assert "YJV22“敷设在土壤中”的数据不用于YJV" in page.text
    assert ">槽盒<" in page.text
    assert 'value="trunking"' not in page.text
    assert "线槽基础载流量依据尚未核实" in page.text
    assert 'name="rcd_scenario"' in page.text
    assert "负荷类型（用于功率因数）" in page.text
    assert "这不是需要系数。需要系数只在“安装功率”模式单独填写。" in page.text
    assert ">2根单芯线<" in page.text
    assert ">3根单芯线<" in page.text
    assert ">三芯电缆<" in page.text
    assert "L1、L2、L3、N各一根" not in page.text

    result = client.post("/quick", data={
        "input_basis": "kva", "input_value": "30", "phase": "3", "voltage_v": "380",
        "load_type_code": "", "conductor_family": "YJV", "installation_scenario": "tray",
    })
    assert result.status_code == 200
    assert "45.5803 A" in result.text
    assert "断路器设计参数要求" in result.text
    assert "微型断路器（MCB）" in result.text
    assert "塑壳断路器（MCCB）" not in result.text
    assert "Icu待算" in result.text
    assert "N/PE" in result.text
    assert "当前结构与已核实的YJV三芯基础载流量表一致" in result.text
    assert "极数：待按接地系统与保护要求确定" in result.text
    assert "框架断路器（ACB）" not in result.text
    assert "该芯数仅用于电缆/导线查表，不决定断路器极数" in result.text
    assert web.db.list_projects() == []
    assert len(web.db.list_rules()) == before_rules


def test_complete_circuit_page_builds_design_and_audit_from_engineering_inputs(tmp_path, monkeypatch):
    monkeypatch.setattr(web, "db", Database(tmp_path / "complete-circuit.db"))
    client = TestClient(web.app)
    page = client.get("/complete-circuit")
    assert page.status_code == 200
    assert "完整低压回路" in page.text
    assert "单电源 · 三相 · 放射式" in page.text
    assert 'name="task_mode"' in page.text
    assert 'name="transformer_capacity_kva"' in page.text
    assert 'name="upstream_short_circuit_capacity_mva"' in page.text
    assert 'name="load_value"' in page.text
    assert 'name="length_connection"' in page.text
    assert 'name="length_feeder"' in page.text
    assert 'name="length_final"' in page.text
    assert 'name="existing_section_connection"' in page.text
    assert "用户不填写专业等值参数" in page.text

    response = client.post("/complete-circuit", data={})
    assert response.status_code == 200
    assert "系统推导的入口参数" in response.text
    assert "50.6448 A" in response.text
    assert "当前主方案" in response.text
    assert "YJV-0.6/1kV" in response.text
    assert "MCCB" in response.text
    assert "逐节点结果" in response.text

    audit = client.post("/complete-circuit", data={"task_mode": "audit"})
    assert audit.status_code == 200
    assert "原设计复核" in audit.text
    assert "QF0 250A" in audit.text
    assert "原电缆结论" in audit.text
    assert "独立替代主方案" in audit.text


def test_complete_circuit_can_be_saved_as_immutable_project_version(tmp_path, monkeypatch):
    database = Database(tmp_path / "network-project.db")
    monkeypatch.setattr(web, "db", database)
    client = TestClient(web.app)
    project_id = database.create_project("P-004", "完整回路归档")

    preview = client.post("/complete-circuit", data={})
    assert f'/projects/{project_id}/complete-circuit' in preview.text
    saved = client.post(
        f"/projects/{project_id}/complete-circuit",
        data={},
        follow_redirects=False,
    )
    assert saved.status_code == 303
    assert saved.headers["location"].startswith("/network-runs/")

    run_id = int(saved.headers["location"].rsplit("/", 1)[1])
    detail = client.get(f"/network-runs/{run_id}")
    assert detail.status_code == 200
    assert "保存时的主方案快照" in detail.text
    assert "完整回路 V1" in detail.text
    assert database.get_project_network(project_id)["revision"] == 1
    assert database.get_network_run(run_id)["stale"] == 0
    pdf = client.get(f"/network-runs/{run_id}/report.pdf")
    assert pdf.status_code == 200
    assert pdf.content.startswith(b"%PDF")
    workbook = client.get(f"/network-runs/{run_id}/export.xlsx")
    assert workbook.status_code == 200
    assert workbook.content.startswith(b"PK")
    exported = load_workbook(BytesIO(workbook.content), data_only=False)
    assert exported.sheetnames == [
        "成果总览", "输入与推导", "电缆方案", "保护器件",
        "逐节点校核", "警告与未闭合项", "依据快照",
    ]
    assert exported["成果总览"]["B6"].value == "V1 / #1"

    changed = client.post(
        f"/projects/{project_id}/complete-circuit",
        data={"load_value": "45"},
        follow_redirects=False,
    )
    assert changed.status_code == 303
    assert database.get_project_network(project_id)["revision"] == 2
    assert database.get_network_run(run_id)["stale"] == 1
    project = client.get(f"/projects/{project_id}")
    assert "完整回路版本" in project.text
    assert "已过期" in project.text

    audit_saved = client.post(
        f"/projects/{project_id}/complete-circuit",
        data={"task_mode": "audit"},
        follow_redirects=False,
    )
    audit_run_id = int(audit_saved.headers["location"].rsplit("/", 1)[1])
    audit_workbook = load_workbook(
        BytesIO(client.get(f"/network-runs/{audit_run_id}/export.xlsx").content),
        data_only=False,
    )
    assert "原设计核验" in audit_workbook.sheetnames
    assert audit_workbook["原设计核验"]["A2"].value == "电缆"


def test_motor_page_uses_exact_reference_row_without_requiring_manual_parameters(tmp_path, monkeypatch):
    monkeypatch.setattr(web, "db", Database(tmp_path / "motor.db"))
    client = TestClient(web.app)

    page = client.get("/motor")
    assert page.status_code == 200
    assert "三相电动机回路" in page.text
    assert 'name="known_value"' in page.text
    assert 'id="motor-catalog-power-selector"' in page.text
    assert '<option value="30.0" selected>30.0 kW · 完整选型</option>' in page.text
    assert '<option value="11.0" >11.0 kW · 完整选型</option>' in page.text
    assert '<option value="200.0" >200.0 kW · 完整选型</option>' in page.text
    assert 'name="motor_efficiency_percent"' in page.text
    assert 'name="motor_power_factor"' in page.text
    assert 'name="starting_frequency"' in page.text
    assert 'name="bus_load_condition"' in page.text
    assert 'name="conductor_configuration"' in page.text
    assert 'name="installation_scenario"' in page.text
    assert 'name="length_m"' in page.text
    assert "不填写修正条件时仍显示基础载流量候选" in page.text
    assert 'name="transformer_family"' in page.text
    assert 'name="upstream_short_circuit_capacity_mva"' in page.text
    assert 'name="motor_starting_time_s"' in page.text
    assert "留空时不执行启动电压近似法" in page.text
    assert '<option value="scb11" selected>SCB11干式变压器</option>' in page.text
    assert 'name="transformer_capacity_kva" type="number" min="1" step="1" value="630"' in page.text
    assert 'name="transformer_uk_percent" type="number" min="0.1" step="0.1" value="6"' in page.text
    assert 'name="installation_temperature_c" type="number" step="1" value="40"' in page.text
    assert "当前采用的快速工况" in page.text

    response = client.post(
        "/motor",
        data={
            "known_basis": "rated_output_power_kw",
            "known_value": "30",
            "rated_voltage_v": "380",
            "poles": "4",
            "starting_frequency": "infrequent",
            "bus_load_condition": "lighting_or_sensitive_loads",
            "conductor_configuration": "yjv_3c_3ph_pe",
            "installation_scenario": "tray",
            "length_m": "50",
        },
    )
    assert response.status_code == 200
    assert "系统自动带出的参考参数" in response.text
    assert "93.6%" in response.text
    assert "0.84" in response.text
    assert "7.3 × In" in response.text
    assert "启动时母线最低电压" in response.text
    assert "≥ 85.0% Un" in response.text
    assert "暂不输出产品型号" in response.text
    assert "电缆基础候选与运行压降" in response.text
    assert "运行压降" in response.text
    assert "缺少修正条件" in response.text
    assert "主保护与控制方案" in response.text
    assert "NXC-65 / 380～400V AC-3" in response.text
    assert "NXR-100 55～70 A" in response.text
    assert "NXR-100 48～65 A" in response.text
    assert "制造商2类配合正式状态：无法判断" in response.text
    assert "制造商2类配合暂算：不适用" in response.text
    assert "该2类配合表只标明IE1/IE2；当前电动机为IE3，不能套用" in response.text
    assert "gG 100A 或 aM 63A" in response.text
    assert "NXM-63H系列 / In 63A" in response.text
    assert "路线A：可调整定MPCB＋接触器" in response.text
    assert "NS2-80/65（48～65A）＋NC8-80" in response.text
    assert "正泰NS2独立产品参数核对" in response.text
    assert "0.8Ii=728.0 A为0.2s不动作试验边界" in response.text
    assert "1.2Ii=1092.0 A为0.2s内动作试验边界" in response.text
    assert "380/400V同一应用电压档" in response.text
    assert "正泰NS2＋NC8制造商2类配合" in response.text
    assert "单品过载、瞬时和分断参数，与厂家短路条件下的2类配合，是两项不同证据" in response.text
    assert "IE3/IE4制造商2类配合参考" in response.text
    assert "3RA2130-4XA37-0AP0" in response.text
    assert "3RA2130-4JA37-0AP0" in response.text
    assert "380/400V同一应用电压档" in response.text
    assert "当前主方案：独立热继电器" in response.text
    assert "NXR-100 48～65 A" in response.text
    assert "首选基础截面" in response.text
    assert "查看其他基础电缆备选" in response.text
    assert "路线B：短路保护器件＋接触器＋独立热继电器" in response.text
    assert "不配置独立NXR" in response.text


def test_motor_default_form_runs_complete_network_without_professional_inputs(
    tmp_path, monkeypatch
):
    monkeypatch.setattr(web, "db", Database(tmp_path / "motor-default.db"))
    client = TestClient(web.app)

    response = client.post("/motor", data=web._motor_form_defaults())

    assert response.status_code == 200
    assert "推荐截面：YJV-0.6/1kV 四芯电缆 25mm²" in response.text
    assert "当前推荐与采购参数" in response.text
    assert "YJV-0.6/1kV 3×25＋1×16 mm² 铜芯电缆" in response.text
    assert "CVS100-MA" in response.text
    assert "LC1E65" in response.text
    assert "有条件采用" in response.text
    assert "采购或调试前必须确认" in response.text


def test_motor_11kw_now_uses_exact_type1_coordination_row(
    tmp_path, monkeypatch
):
    monkeypatch.setattr(web, "db", Database(tmp_path / "motor-limited.db"))
    client = TestClient(web.app)
    form = web._motor_form_defaults()
    form["known_value"] = "11"

    response = client.post("/motor", data=form)

    assert response.status_code == 200
    assert "当前推荐与采购参数" in response.text
    assert "CVS100-MA" in response.text
    assert "LC1E25" in response.text
    assert "LRE22" in response.text
    assert "额定运行电流" in response.text


def test_motor_200kw_page_uses_high_power_catalog_and_exact_type1_devices(
    tmp_path, monkeypatch
):
    monkeypatch.setattr(web, "db", Database(tmp_path / "motor-200kw.db"))
    client = TestClient(web.app)
    form = web._motor_form_defaults()
    form["known_value"] = "200"

    response = client.post("/motor", data=form)

    assert response.status_code == 200
    assert "1LE1503" in response.text
    assert "3/20" in response.text
    assert "PDF第158页" in response.text
    assert "CVS630-MA" in response.text
    assert "LC1F400" in response.text
    assert "LR9-F7379" in response.text
    assert "当前推荐与采购参数" in response.text


def test_motor_page_does_not_interpolate_missing_reference_power(tmp_path, monkeypatch):
    monkeypatch.setattr(web, "db", Database(tmp_path / "motor-missing.db"))
    client = TestClient(web.app)

    response = client.post(
        "/motor",
        data={
            "known_basis": "rated_output_power_kw",
            "known_value": "20",
            "rated_voltage_v": "380",
            "poles": "4",
            "starting_frequency": "infrequent",
            "bus_load_condition": "lighting_or_sensitive_loads",
            "conductor_configuration": "yjv_3c_3ph_pe",
            "installation_scenario": "tray",
            "length_m": "50",
            "transformer_family": "scb11",
            "transformer_capacity_kva": "630",
            "transformer_uk_percent": "6",
        },
    )
    assert response.status_code == 200
    assert "没有精确参考行" in response.text
    assert "系统不插值" in response.text
    assert "补充效率和运行功率因数" in response.text
    assert "运行电流</strong><small>缺参数" in response.text
    assert "启动与短路</strong><small>被上游参数阻断" in response.text


def test_motor_page_redirects_unsupported_four_core_conduit_to_verified_three_core_path(
    tmp_path, monkeypatch
):
    monkeypatch.setattr(web, "db", Database(tmp_path / "motor-cable-path.db"))
    client = TestClient(web.app)

    response = client.post(
        "/motor",
        data={
            "known_basis": "rated_output_power_kw",
            "known_value": "30",
            "rated_voltage_v": "380",
            "poles": "4",
            "starting_frequency": "infrequent",
            "bus_load_condition": "no_lighting_or_sensitive_loads",
            "conductor_configuration": "yjv_4c_3ph_n_pe",
            "installation_scenario": "conduit",
            "length_m": "50",
        },
    )

    assert response.status_code == 200
    assert "系统已改用有已核实表格的YJV三芯电缆＋独立PE" in response.text
    assert "首选基础截面：YJV-0.6/1kV 三芯电缆" in response.text
    assert "没有电缆候选" not in response.text
    assert "长延时规范关系" in response.text
    assert "原依据没有给出“接近”的百分比上限" in response.text
    assert "瞬时整定明确范围" in response.text
    assert "规范可调范围要求" in response.text
    assert "所选产品可调范围：</strong>48～65 A" in response.text


def test_motor_page_accepts_verified_manual_parameters_when_catalog_has_no_exact_row(
    tmp_path, monkeypatch
):
    monkeypatch.setattr(web, "db", Database(tmp_path / "motor-manual.db"))
    client = TestClient(web.app)

    response = client.post(
        "/motor",
        data={
            "known_basis": "rated_output_power_kw",
            "known_value": "50",
            "rated_voltage_v": "380",
            "poles": "4",
            "motor_efficiency_percent": "94",
            "motor_power_factor": "0.85",
            "locked_rotor_current_ratio": "7",
            "starting_frequency": "infrequent",
            "bus_load_condition": "no_lighting_or_sensitive_loads",
            "conductor_configuration": "yjv_3c_3ph_pe",
            "installation_scenario": "conduit",
            "length_m": "100",
            "installation_temperature_c": "35",
            "enclosed_circuit_count": "1",
            "transformer_family": "scb11",
            "transformer_capacity_kva": "630",
            "transformer_uk_percent": "6",
            "upstream_short_circuit_capacity_mva": "100",
        },
    )

    assert response.status_code == 200
    assert "用户补充的铭牌或厂家参数" in response.text
    assert "94.0%" in response.text
    assert "0.85" in response.text
    assert "95.078" in response.text
    assert "推荐电缆完整网络复核" in response.text
    assert "启动与短路</strong><small>已执行，部分缺参数" in response.text
    assert "需接入变压器—线路完整网络" not in response.text


def test_motor_page_rejects_partial_or_invalid_manual_power_parameters(
    tmp_path, monkeypatch
):
    monkeypatch.setattr(web, "db", Database(tmp_path / "motor-manual-invalid.db"))
    client = TestClient(web.app)

    partial = client.post(
        "/motor",
        data={
            "known_basis": "rated_output_power_kw",
            "known_value": "50",
            "rated_voltage_v": "380",
            "poles": "4",
            "motor_efficiency_percent": "94",
            "motor_power_factor": "",
            "starting_frequency": "infrequent",
            "bus_load_condition": "no_lighting_or_sensitive_loads",
            "conductor_configuration": "yjv_3c_3ph_pe",
            "installation_scenario": "conduit",
            "length_m": "100",
        },
    )
    assert "效率和运行功率因数必须同时填写" in partial.text

    invalid = client.post(
        "/motor",
        data={
            "known_basis": "rated_output_power_kw",
            "known_value": "50",
            "rated_voltage_v": "380",
            "poles": "4",
            "motor_efficiency_percent": "105",
            "motor_power_factor": "1.2",
            "starting_frequency": "infrequent",
            "bus_load_condition": "no_lighting_or_sensitive_loads",
            "conductor_configuration": "yjv_3c_3ph_pe",
            "installation_scenario": "conduit",
            "length_m": "100",
        },
    )
    assert "效率必须大于0且不大于100%" in invalid.text
    assert "运行功率因数必须大于0且不大于1" in invalid.text


def test_motor_page_can_recalculate_each_cable_in_complete_network(tmp_path, monkeypatch):
    monkeypatch.setattr(web, "db", Database(tmp_path / "motor-network.db"))
    client = TestClient(web.app)

    response = client.post(
        "/motor",
        data={
            "known_basis": "rated_output_power_kw",
            "known_value": "30",
            "rated_voltage_v": "380",
            "poles": "4",
            "starting_frequency": "infrequent",
            "bus_load_condition": "lighting_or_sensitive_loads",
            "conductor_configuration": "yjv_4c_3ph_n_pe",
            "installation_scenario": "tray",
            "length_m": "50",
            "installation_temperature_c": "40",
            "tray_layers": "1",
            "tray_cables_per_layer": "1",
            "transformer_family": "scb11",
            "transformer_capacity_kva": "630",
            "transformer_uk_percent": "6",
            "upstream_short_circuit_capacity_mva": "100",
            "preconnected_reactive_load_mvar": "0.1",
        },
    )

    assert response.status_code == 200
    assert "推荐电缆完整网络复核" in response.text
    assert "已形成推荐设计截面" in response.text
    assert "推荐：YJV-0.6/1kV 四芯电缆 25mm²" in response.text
    assert "当前推荐与采购参数" in response.text
    assert "查看计算过程、其他方案和专业校核" in response.text
    assert response.text.index("当前推荐与采购参数") < response.text.index(
        "电流与启动要求"
    )
    assert "YJV-0.6/1kV 3×25＋1×16 mm² 铜芯电缆" in response.text
    assert "CVS100-MA" in response.text
    assert "Irm 900A" in response.text
    assert "LRE359" in response.text
    assert "有条件采用" in response.text
    assert "制造商IEC/EN 60947-4-1、380～415V直接启动1类配合表精确行" in response.text
    assert "跨品牌组合不等于取得制造商1类/2类配合认证" in response.text
    assert "正泰：可调整定MPCB＋AC-3接触器" in response.text
    assert "德力西保护器＋正泰控制器：电子式MCCB短路保护＋AC-3接触器＋独立热继电器" in response.text
    assert "未成为主方案的原因" in response.text
    assert "采购或调试前必须确认" in response.text
    assert "电动机保护型断路器已经承担过载保护时，不重复配置独立热继电器" in response.text
    assert "专业深化项" in response.text
    assert "允许整定区间：</strong>57.97～113.0 A" in response.text
    assert "末端最大三相短路" in response.text
    assert "末端最小相—PE故障" in response.text
    assert "相导体允许最长切除" in response.text
    assert "启动母线/端子电压" in response.text
    assert "该电缆对应的断路器必须满足" in response.text
    assert "Icu：≥" in response.text
    assert "通用参数档位" in response.text
    assert "不能仅凭瞬时段证明，须查实际反时限曲线" in response.text
    assert "新版同类产品参考（不是品牌推荐）" in response.text
    assert "CM3-63L 电动机保护型" in response.text
    assert "瞬时整定：12In" in response.text
    assert "曲线采用保守栅格数字化边界" in response.text
    assert "过载基准：固定In 63.0 A" in response.text
    assert "产品保证点：1.0In=63.0 A冷态2h内不动作" in response.text
    assert "与电动机允许过载特性匹配：无法判断" in response.text
    assert "安装点相间故障" in response.text
    assert "末端相—PE故障" in response.text
    assert "综合判定：</strong>通过" in response.text
    assert "正泰NS2逐电缆候选联动（不是品牌推荐）" in response.text
    assert "进入保证瞬时动作区：" in response.text
    assert "NS2完整切除时间校核：无法判断" in response.text
    assert "未明确为包含触头完全开断的总切除时间" in response.text
    assert "说明书Figure 1虽提供20℃时间-电流特性曲线" in response.text


def test_motor_page_exact_400v_can_render_siemens_ie3_type2_evidence(tmp_path, monkeypatch):
    monkeypatch.setattr(web, "db", Database(tmp_path / "motor-400v.db"))
    client = TestClient(web.app)

    response = client.post(
        "/motor",
        data={
            "known_basis": "rated_output_power_kw",
            "known_value": "30",
            "rated_voltage_v": "400",
            "poles": "4",
            "starting_frequency": "infrequent",
            "bus_load_condition": "lighting_or_sensitive_loads",
            "conductor_configuration": "yjv_4c_3ph_n_pe",
            "installation_scenario": "tray",
            "length_m": "50",
            "installation_temperature_c": "40",
            "tray_layers": "1",
            "tray_cables_per_layer": "1",
            "transformer_family": "scb11",
            "transformer_capacity_kva": "630",
            "transformer_uk_percent": "6",
            "upstream_short_circuit_capacity_mva": "100",
            "preconnected_reactive_load_mvar": "0.1",
        },
    )

    assert response.status_code == 200
    assert "IE3/IE4制造商2类配合参考" in response.text
    assert "3RA2130-4XA37-0AP0" in response.text
    assert "3RA2130-4JA37-0AP0" in response.text
    assert "不超过表列Iq 100kA" in response.text
    assert "暂算：通过" in response.text
    assert "运行压降：无法判断" not in response.text
    assert "启动电流≤720 A（通过）" in response.text
    assert "单品Icu 65 kA（通过）" in response.text
    assert "完整动作曲线与最终配合证据保留在专业复核中" in response.text
    assert "推荐：YJV-0.6/1kV 四芯电缆 25mm²" in response.text


def test_quick_transformer_lv_outlet_short_circuit_lookup(tmp_path, monkeypatch):
    monkeypatch.setattr(web, "db", Database(tmp_path / "quick-short-circuit.db"))
    client = TestClient(web.app)

    result = client.post("/quick", data={
        "input_basis": "current", "input_value": "100", "phase": "3", "voltage_v": "380",
        "conductor_family": "BV", "installation_scenario": "conduit",
        "short_circuit_method": "transformer_lv_table",
        "transformer_capacity_kva": "1000", "transformer_uk_percent": "6",
    })

    assert result.status_code == 200
    assert "变压器0.4kV出口短路暂算" in result.text
    assert "24.0 kA" in result.text
    assert "61.2 kA" in result.text
    assert "表15.7" in result.text
    assert "Icu 要求超出当前表列档位" in result.text
    assert "塑壳断路器（MCCB）" not in result.text


def test_quick_unknown_kw_accepts_user_power_factor(tmp_path, monkeypatch):
    monkeypatch.setattr(web, "db", Database(tmp_path / "quick-pf.db"))
    client = TestClient(web.app)

    result = client.post("/quick", data={
        "input_basis": "kw", "input_value": "10", "phase": "1", "voltage_v": "220",
        "load_type_code": "unknown", "power_factor": "0.8",
        "conductor_family": "BV", "installation_scenario": "conduit",
    })

    assert result.status_code == 200
    assert "56.8182 A" in result.text
    assert "用户输入（应与铭牌或厂家资料核对）" in result.text
    assert "正式使用前应与设备铭牌或厂家资料核对" in result.text


def test_quick_line_end_short_circuit_is_opt_in_and_does_not_change_old_flow(tmp_path, monkeypatch):
    monkeypatch.setattr(web, "db", Database(tmp_path / "quick-line-disabled.db"))
    captured = capture_template_context(monkeypatch)
    client = TestClient(web.app)

    response = client.post("/quick", data={
        "input_basis": "current", "input_value": "100", "phase": "3", "voltage_v": "380",
        "conductor_family": "YJV", "installation_scenario": "conduit",
        "length_m": "100", "transformer_capacity_kva": "1000", "transformer_uk_percent": "6",
    })

    assert response.status_code == 200
    assert "line_end_short_circuit" not in captured["result"]


def test_quick_line_end_short_circuit_uses_verified_nameplate_rx_case(tmp_path, monkeypatch):
    monkeypatch.setattr(web, "db", Database(tmp_path / "quick-line-complete.db"))
    captured = capture_template_context(monkeypatch)
    client = TestClient(web.app)

    response = client.post("/quick", data={
        "input_basis": "current", "input_value": "100", "phase": "3", "voltage_v": "400",
        "conductor_family": "YJV", "installation_scenario": "conduit", "length_m": "100",
        "short_circuit_line_enabled": "on",
        "short_circuit_line_type": "cable", "short_circuit_line_section_mm2": "50",
        "transformer_capacity_kva": "1000", "transformer_uk_percent": "6",
        "transformer_pk_kw": "10",
        "source_impedance_mode": "infinite_capacity",
        "voltage_factor_c": "1.05",
        "breaker_installation_point": "line_end",
        "breaker_icu_ka": "25",
    })

    line_result = captured["result"]["line_end_short_circuit"]
    assert response.status_code == 200
    assert line_result["outputs"]["terminal_short_circuit_current_ka"] > 0
    assert line_result["outputs"]["upstream_impedance"]["method"].startswith("变压器铭牌 R/X")
    assert line_result["status"] == "无法判断"


def test_quick_line_end_short_circuit_converts_source_short_circuit_capacity(tmp_path, monkeypatch):
    monkeypatch.setattr(web, "db", Database(tmp_path / "quick-line-source-capacity.db"))
    captured = capture_template_context(monkeypatch)
    client = TestClient(web.app)

    response = client.post("/quick", data={
        "input_basis": "current", "input_value": "100", "phase": "3", "voltage_v": "400",
        "conductor_family": "YJV", "installation_scenario": "conduit", "length_m": "100",
        "short_circuit_line_enabled": "on", "short_circuit_line_type": "cable",
        "transformer_capacity_kva": "1000", "transformer_uk_percent": "6", "transformer_pk_kw": "10",
        "source_impedance_mode": "short_circuit_capacity", "source_short_circuit_capacity_mva": "100",
        "voltage_factor_c": "1.05", "breaker_installation_point": "line_end", "breaker_icu_ka": "25",
    })

    line_result = captured["result"]["line_end_short_circuit"]
    source = line_result["outputs"]["upstream_impedance"]["source_equivalent"]
    assert response.status_code == 200
    assert source["short_circuit_capacity_mva"] == 100
    assert source["impedance_ohm"] == pytest.approx(0.0016)


def test_quick_shared_transformer_fields_derive_rx_and_icu_requirement(tmp_path, monkeypatch):
    monkeypatch.setattr(web, "db", Database(tmp_path / "quick-shared-transformer.db"))
    captured = capture_template_context(monkeypatch)
    client = TestClient(web.app)

    response = client.post("/quick", data={
        "input_basis": "current", "input_value": "100", "phase": "3", "voltage_v": "380",
        "conductor_family": "YJV", "conductor_configuration": "yjv_5c_3ph_n_pe",
        "installation_scenario": "tray", "tray_type": "horizontal_perforated",
        "tray_layers": "1", "tray_cables_per_layer": "1",
        "installation_temperature_c": "40", "length_m": "100",
        "short_circuit_line_enabled": "on", "short_circuit_line_type": "cable",
        "transformer_series_code": "scb11", "transformer_capacity_kva": "630",
        "transformer_uk_percent": "6", "source_impedance_mode": "short_circuit_capacity",
        "source_short_circuit_capacity_mva": "100",
    })

    line_result = captured["result"]["line_end_short_circuit"]
    transformer = line_result["outputs"]["upstream_impedance"]["transformer_equivalent"]
    assert response.status_code == 200
    assert transformer["resistance_ohm"] == pytest.approx(0.0024)
    assert transformer["reactance_ohm"] == pytest.approx(0.015)
    assert line_result["outputs"]["required_breaking_capacity_ka"] > 0
    assert line_result["outputs"]["breaker_icu_check"]["note"].startswith("设计模式")
    short_stage = next(
        item for item in captured["result"]["outputs"]["workflow_stages"]
        if item["code"] == "short_circuit"
    )
    assert short_stage == {
        "code": "short_circuit", "label": "短路电流与Icu要求", "state": "completed"
    }
    assert "已选断路器Icu实物复核" in captured["result"]["outputs"]["incomplete_checks"]
    assert not any(
        warning.startswith("未提供安装点预期短路电流")
        for warning in captured["result"]["warnings"]
    )


def test_quick_line_end_short_circuit_missing_nameplate_or_source_stays_unknown(tmp_path, monkeypatch):
    monkeypatch.setattr(web, "db", Database(tmp_path / "quick-line-missing.db"))
    captured = capture_template_context(monkeypatch)
    client = TestClient(web.app)
    base = {
        "input_basis": "current", "input_value": "100", "phase": "3", "voltage_v": "400",
        "conductor_family": "YJV", "installation_scenario": "conduit", "length_m": "100",
        "short_circuit_line_enabled": "on",
        "short_circuit_line_type": "cable", "short_circuit_line_section_mm2": "50",
        "transformer_capacity_kva": "1000", "transformer_uk_percent": "6",
        "voltage_factor_c": "1.05",
    }

    missing_pk = client.post("/quick", data={
        **base, "source_impedance_mode": "infinite_capacity",
    })
    missing_pk_result = captured["result"]["line_end_short_circuit"]
    assert missing_pk.status_code == 200
    assert missing_pk_result["status"] == "无法判断"
    assert any("transformer_pk_kw" in item for item in missing_pk_result["warnings"])

    missing_source = client.post("/quick", data={
        **base, "transformer_pk_kw": "10",
    })
    missing_source_result = captured["result"]["line_end_short_circuit"]
    assert missing_source.status_code == 200
    assert missing_source_result["status"] == "无法判断"
    assert any("未明确上级系统无限容量" in item for item in missing_source_result["warnings"])


def test_quick_line_end_short_circuit_maps_template_busway_fields(tmp_path, monkeypatch):
    monkeypatch.setattr(web, "db", Database(tmp_path / "quick-line-busway.db"))
    captured = capture_template_context(monkeypatch)
    client = TestClient(web.app)

    response = client.post("/quick", data={
        "input_basis": "current", "input_value": "100", "phase": "3", "voltage_v": "400",
        "conductor_family": "YJV", "installation_scenario": "tray", "length_m": "50",
        "short_circuit_line_enabled": "true", "short_circuit_line_type": "busway",
        "upstream_r_ohm": "0.003", "upstream_x_ohm": "0.020",
        "upstream_impedance_reference": "上游短路计算书第1页", "voltage_factor_c": "1.05",
        "busway_r_ohm_per_km": "0.020", "busway_x_ohm_per_km": "0.015",
        "busway_impedance_reference": "制造商样本第12页",
    })

    line_result = captured["result"]["line_end_short_circuit"]
    assert response.status_code == 200
    assert line_result["outputs"]["line_impedance"]["source"]["mode"] == "explicit"
    assert line_result["outputs"]["terminal_short_circuit_current_ka"] > 0


def test_quick_cable_does_not_reuse_submitted_busway_rx(tmp_path, monkeypatch):
    monkeypatch.setattr(web, "db", Database(tmp_path / "quick-line-cable.db"))
    captured = capture_template_context(monkeypatch)
    client = TestClient(web.app)

    response = client.post("/quick", data={
        "input_basis": "current", "input_value": "100", "phase": "3", "voltage_v": "400",
        "conductor_family": "YJV", "installation_scenario": "conduit", "length_m": "50",
        "short_circuit_line_enabled": "true", "short_circuit_line_type": "cable",
        "short_circuit_line_section_mm2": "50",
        "upstream_r_ohm": "0.003", "upstream_x_ohm": "0.020",
        "upstream_impedance_reference": "上游短路计算书第1页", "voltage_factor_c": "1.05",
        "busway_r_ohm_per_km": "0.020", "busway_x_ohm_per_km": "0.015",
        "busway_impedance_reference": "不应被采用",
    })

    line_result = captured["result"]["line_end_short_circuit"]
    assert response.status_code == 200
    assert line_result["outputs"]["line_impedance"]["source"]["table"] == "表3.21"
    assert line_result["outputs"]["line_impedance"]["source"].get("mode") != "explicit"


def test_quick_earth_fault_is_opt_in(tmp_path, monkeypatch):
    monkeypatch.setattr(web, "db", Database(tmp_path / "quick-earth-disabled.db"))
    captured = capture_template_context(monkeypatch)
    client = TestClient(web.app)

    response = client.post("/quick", data={
        "input_basis": "current", "input_value": "32", "phase": "1", "voltage_v": "220",
        "conductor_family": "BV", "installation_scenario": "conduit",
    })

    assert response.status_code == 200
    assert "earth_fault" not in captured["result"]


def test_quick_earth_fault_maps_traceable_zs_and_ia(tmp_path, monkeypatch):
    monkeypatch.setattr(web, "db", Database(tmp_path / "quick-earth-complete.db"))
    captured = capture_template_context(monkeypatch)
    client = TestClient(web.app)

    response = client.post("/quick", data={
        "input_basis": "current", "input_value": "32", "phase": "1", "voltage_v": "220",
        "conductor_family": "BV", "installation_scenario": "conduit",
        "earth_fault_enabled": "true", "earthing_system": "TN-S",
        "nominal_line_to_earth_voltage_v": "230", "circuit_application": "socket_final",
        "circuit_rated_current_a": "32", "fault_loop_impedance_ohm": "0.8",
        "fault_loop_impedance_reference": "接地故障回路计算书第2页",
        "protection_type": "overcurrent", "protective_device_operating_current_a": "200",
        "protective_device_operating_reference": "断路器时间—电流曲线第3页，0.4s",
    })

    earth = captured["result"]["earth_fault"]
    assert response.status_code == 200
    assert earth["outputs"]["maximum_disconnection_time_s"] == 0.4
    assert earth["outputs"]["prospective_earth_fault_current_a"] == 275
    assert earth["provisional_status"] == "通过"
    assert earth["status"] == "无法判断"
    assert "接地故障与自动切断暂算" in response.text


def test_quick_four_core_renders_multicore_ampacity_and_structure_source(tmp_path, monkeypatch):
    monkeypatch.setattr(web, "db", Database(tmp_path / "quick-yjv-four-core.db"))
    client = TestClient(web.app)

    response = client.post("/quick", data={
        "input_basis": "current", "input_value": "50", "phase": "3", "voltage_v": "380",
        "conductor_family": "YJV",
        "conductor_configuration": "yjv_4c_3ph_n_pe",
        "installation_scenario": "tray",
        "tray_type": "horizontal_perforated",
        "tray_layers": "1",
        "tray_cables_per_layer": "1",
    })

    assert response.status_code == 200
    assert "表31" in response.text
    assert "推荐最小截面：YJV 10.0 mm²" in response.text
    assert "表列中性线（接地线） 6 mm²" in response.text


def test_quick_short_circuit_uses_selected_cable_section_and_derives_u0(
    tmp_path, monkeypatch
):
    monkeypatch.setattr(web, "db", Database(tmp_path / "quick-derived-short-circuit.db"))
    captured = capture_template_context(monkeypatch)
    client = TestClient(web.app)

    response = client.post("/quick", data={
        "input_basis": "current", "input_value": "50", "phase": "3", "voltage_v": "380",
        "conductor_family": "YJV", "installation_scenario": "conduit", "length_m": "50",
        "short_circuit_line_enabled": "true", "short_circuit_line_type": "cable",
        "transformer_capacity_kva": "1000", "transformer_uk_percent": "6",
        "transformer_pk_kw": "10", "source_impedance_mode": "infinite_capacity",
        "voltage_factor_c": "1.05", "breaker_installation_point": "line_end",
        "breaker_icu_ka": "25",
        "earth_fault_enabled": "true", "earthing_system": "TN-S",
        "circuit_application": "distribution", "protection_type": "overcurrent",
        "protective_device_characteristic": "mcb_c",
    })

    result = captured["result"]
    selection = result["line_short_circuit_selection"]
    earth = result["earth_fault"]
    candidate = result["outputs"]["cable_candidates"][0]
    assert response.status_code == 200
    assert selection["section_mm2"] == str(candidate["section_mm2"])
    assert selection["source"] == "采用本次电缆初选后的最终截面"
    assert earth["outputs"]["nominal_line_to_earth_voltage_v"] == 220
    assert earth["outputs"]["nominal_line_to_earth_voltage_source"] == (
        "三相380V系统：U0自动采用220V。"
    )


def test_quick_earth_fault_uses_selected_yjv_conventional_method(tmp_path, monkeypatch):
    monkeypatch.setattr(web, "db", Database(tmp_path / "quick-yjv-earth.db"))
    captured = capture_template_context(monkeypatch)
    client = TestClient(web.app)

    response = client.post("/quick", data={
        "input_basis": "current", "input_value": "50", "phase": "3", "voltage_v": "380",
        "conductor_family": "YJV",
        "conductor_configuration": "yjv_4c_3ph_n_pe",
        "installation_scenario": "tray",
        "tray_type": "horizontal_perforated",
        "tray_layers": "1",
        "tray_cables_per_layer": "1",
        "length_m": "50",
        "earth_fault_enabled": "true",
        "earthing_system": "TN-S",
        "nominal_line_to_earth_voltage_v": "230",
        "circuit_application": "distribution",
        "protection_type": "overcurrent",
        "protective_device_characteristic": "mcb_c",
        "fault_fourth_conductor_role": "PE",
    })

    earth = captured["result"]["earth_fault"]
    assert response.status_code == 200
    assert earth["outputs"]["fault_current_calculation_method"] == "tn_conventional"
    assert earth["outputs"]["conventional_method"]["phase_section_mm2"] == 10
    assert earth["outputs"]["conventional_method"]["protective_section_mm2"] == 6
    assert earth["outputs"]["protective_device_operating_current_a"] == 500
    assert "TN常规法" in response.text


def test_quick_earth_fault_builds_transformer_to_yjv_complete_loop(tmp_path, monkeypatch):
    monkeypatch.setattr(web, "db", Database(tmp_path / "quick-yjv-complete-loop.db"))
    captured = capture_template_context(monkeypatch)
    client = TestClient(web.app)

    response = client.post("/quick", data={
        "input_basis": "current", "input_value": "50", "phase": "3", "voltage_v": "380",
        "conductor_family": "YJV",
        "conductor_configuration": "yjv_4c_3ph_n_pe",
        "installation_scenario": "tray",
        "tray_type": "horizontal_perforated",
        "tray_layers": "1",
        "tray_cables_per_layer": "1",
        "length_m": "50",
        "earth_fault_enabled": "true",
        "earthing_system": "TN-S",
        "nominal_line_to_earth_voltage_v": "230",
        "circuit_application": "distribution",
        "protection_type": "overcurrent",
        "protective_device_characteristic": "mcb_c",
        "fault_transformer_series_code": "s11_m",
        "fault_transformer_capacity_kva": "400",
        "fault_transformer_uk_percent": "4",
        "fault_transformer_hv_voltage_kv": "10",
        "fault_transformer_vector_group": "Dyn11",
        "fault_connection_type": "direct",
        "fault_fourth_conductor_role": "PE",
    })

    earth = captured["result"]["earth_fault"]
    chain = earth["outputs"]["calculated_fault_loop_chain"]["outputs"]
    cable = chain["components"][1]
    assert response.status_code == 200
    assert earth["outputs"]["fault_current_calculation_method"] == "complete_loop_impedance"
    assert chain["target_point"] == "line_end"
    assert cable["cable_specification"] == "YJV-0.6/1kV 3×10+1×6"
    assert cable["length_m"] == 50
    assert cable["phase_pe_resistance_multiplier"] == 1.5
    assert earth["outputs"]["prospective_earth_fault_current_a"] > 0
    assert "完整回路 R / X / Zs" in response.text


def test_quick_complete_loop_looks_up_canalis_busway_phase_pe_rx(
    tmp_path, monkeypatch
):
    monkeypatch.setattr(web, "db", Database(tmp_path / "quick-canalis-loop.db"))
    captured = capture_template_context(monkeypatch)
    client = TestClient(web.app)

    response = client.post("/quick", data={
        "input_basis": "current", "input_value": "50", "phase": "3", "voltage_v": "380",
        "conductor_family": "YJV",
        "conductor_configuration": "yjv_4c_3ph_n_pe",
        "installation_scenario": "tray",
        "tray_type": "horizontal_perforated",
        "tray_layers": "1",
        "tray_cables_per_layer": "1",
        "length_m": "50",
        "earth_fault_enabled": "true",
        "earthing_system": "TN-S",
        "nominal_line_to_earth_voltage_v": "230",
        "circuit_application": "distribution",
        "protection_type": "overcurrent",
        "protective_device_characteristic": "mcb_c",
        "fault_transformer_series_code": "s11_m",
        "fault_transformer_capacity_kva": "400",
        "fault_transformer_uk_percent": "4",
        "fault_transformer_hv_voltage_kv": "10",
        "fault_transformer_vector_group": "Dyn11",
        "fault_connection_type": "busway",
        "fault_busway_series_code": "canalis_kta_casing_pe",
        "fault_busway_rating_a": "1600",
        "fault_connection_length_m": "5",
        "fault_fourth_conductor_role": "PE",
    })

    selection = captured["result"]["busway_phase_pe_selection"]
    earth = captured["result"]["earth_fault"]
    chain = earth["outputs"]["calculated_fault_loop_chain"]["outputs"]
    busway = chain["components"][1]
    assert response.status_code == 200
    assert selection["series_name"] == "Canalis KTA（外壳作PE）"
    assert busway["resistance_ohm_per_km"] == pytest.approx(0.394)
    assert busway["reactance_ohm_per_km"] == pytest.approx(0.212)
    assert busway["resistance_ohm"] == pytest.approx(0.00197)
    assert busway["reactance_ohm"] == pytest.approx(0.00106)
    assert "ELEC.BUSWAY.CANALIS.PHASE_PE.IMPEDANCE" in (
        earth["outputs"]["calculated_fault_loop_chain"]["rule_codes"]
    )
    assert "母线槽相—PE参数" in response.text
    assert "DEBU021EN" in response.text


def test_quick_five_core_builds_complete_loop_from_verified_geometry(
    tmp_path, monkeypatch
):
    monkeypatch.setattr(web, "db", Database(tmp_path / "quick-five-core-loop.db"))
    captured = capture_template_context(monkeypatch)
    client = TestClient(web.app)

    response = client.post("/quick", data={
        "input_basis": "current", "input_value": "50", "phase": "3", "voltage_v": "380",
        "conductor_family": "YJV",
        "conductor_configuration": "yjv_5c_3ph_n_pe",
        "installation_scenario": "tray",
        "tray_type": "horizontal_perforated",
        "tray_layers": "1",
        "tray_cables_per_layer": "1",
        "length_m": "50",
        "earth_fault_enabled": "true",
        "earthing_system": "TN-S",
        "nominal_line_to_earth_voltage_v": "230",
        "circuit_application": "distribution",
        "protection_type": "overcurrent",
        "protective_device_characteristic": "mcb_c",
        "fault_transformer_series_code": "s11_m",
        "fault_transformer_capacity_kva": "400",
        "fault_transformer_uk_percent": "4",
        "fault_transformer_hv_voltage_kv": "10",
        "fault_transformer_vector_group": "Dyn11",
        "fault_connection_type": "direct",
        "fault_fourth_conductor_role": "PE",
    })

    earth = captured["result"]["earth_fault"]
    chain = earth["outputs"]["calculated_fault_loop_chain"]["outputs"]
    cable = chain["components"][1]
    cable_calculation = cable["cable_calculation"]["outputs"]
    assert response.status_code == 200
    assert earth["outputs"]["fault_current_calculation_method"] == (
        "complete_loop_impedance"
    )
    assert cable["name"] == "线路末端YJV五芯回路"
    assert cable_calculation["structure_catalog"]["profile"] == "yjv_3plus2"
    assert cable_calculation["fault_resistance_multiplier"] == 1.5
    assert cable_calculation["reactance_method"] == "geometry"
    assert earth["outputs"]["prospective_earth_fault_current_a"] > 0


def test_quick_five_core_auto_uses_one_core_as_pe_for_design_selection(
    tmp_path, monkeypatch
):
    monkeypatch.setattr(web, "db", Database(tmp_path / "quick-unconfirmed-pe.db"))
    captured = capture_template_context(monkeypatch)
    client = TestClient(web.app)

    response = client.post("/quick", data={
        "input_basis": "current", "input_value": "50", "phase": "3", "voltage_v": "380",
        "conductor_family": "YJV",
        "conductor_configuration": "yjv_5c_3ph_n_pe",
        "installation_scenario": "tray",
        "tray_type": "horizontal_perforated",
        "tray_layers": "1",
        "tray_cables_per_layer": "1",
        "length_m": "50",
        "earth_fault_enabled": "true",
        "earthing_system": "TN-S",
        "nominal_line_to_earth_voltage_v": "230",
        "circuit_application": "distribution",
        "protection_type": "overcurrent",
        "fault_transformer_series_code": "s11_m",
        "fault_transformer_capacity_kva": "400",
        "fault_transformer_uk_percent": "4",
        "fault_transformer_hv_voltage_kv": "10",
        "fault_transformer_vector_group": "Dyn11",
    })

    earth = captured["result"]["earth_fault"]
    stages = captured["result"]["outputs"]["workflow_stages"]
    earth_stage = next(stage for stage in stages if stage["code"] == "earth_fault")
    curve_candidates = earth["outputs"]["protective_device_curve_candidates"]
    assert response.status_code == 200
    assert captured["form"]["fault_fourth_conductor_role"] == "PE"
    assert earth["outputs"]["protective_core_role"] == "PE"
    assert "五芯电缆设计初选" in earth["outputs"]["protective_core_role_source"]
    assert "calculated_fault_loop_chain" in earth["outputs"]
    assert earth["outputs"]["fault_current_calculation_method"] == (
        "complete_loop_impedance"
    )
    assert earth["outputs"]["prospective_earth_fault_current_a"] > 0
    assert earth["outputs"]["maximum_permitted_operating_current_a"] > 0
    assert earth["provisional_status"] == "无法判断"
    assert [candidate["curve"] for candidate in curve_candidates] == ["B", "C"]
    assert all(
        candidate["operating_current_a"] > 0 for candidate in curve_candidates
    )
    assert earth_stage["label"] == "接地故障与保护约束"
    assert earth_stage["state"] == "candidate"
    assert "保护器件曲线/整定实物复核" in (
        captured["result"]["outputs"]["incomplete_checks"]
    )
    assert "参数候选暂算" in response.text


def test_quick_pe_thermal_uses_selected_yjv_pe_and_calculated_fault_current(
    tmp_path, monkeypatch
):
    monkeypatch.setattr(web, "db", Database(tmp_path / "quick-pe-thermal.db"))
    captured = capture_template_context(monkeypatch)
    client = TestClient(web.app)

    response = client.post("/quick", data={
        "input_basis": "current", "input_value": "50", "phase": "3", "voltage_v": "380",
        "conductor_family": "YJV",
        "conductor_configuration": "yjv_4c_3ph_n_pe",
        "installation_scenario": "tray",
        "tray_type": "horizontal_perforated",
        "tray_layers": "1",
        "tray_cables_per_layer": "1",
        "length_m": "50",
        "earth_fault_enabled": "true",
        "earthing_system": "TN-S",
        "nominal_line_to_earth_voltage_v": "230",
        "circuit_application": "distribution",
        "protection_type": "overcurrent",
        "protective_device_characteristic": "mcb_c",
        "fault_fourth_conductor_role": "PE",
        "pe_thermal_enabled": "true",
        "protective_core_confirmed": "true",
        "fault_clearing_time_s": "0.1",
    })

    pe = captured["result"]["pe_thermal"]
    assert response.status_code == 200
    assert pe["outputs"]["protective_conductor_section_mm2"] == 6
    assert pe["outputs"]["k_a_sqrt_s_per_mm2"] == 143
    assert pe["outputs"]["actual_thermal_stress_a2s"] > 0
    assert pe["provisional_status"] == "通过"
    assert pe["status"] == "无法判断"
    assert "PE导体短路热稳定暂算" in response.text
    assert "PE导体热稳定" not in captured["result"]["outputs"]["incomplete_checks"]


def test_quick_pe_thermal_reuses_confirmed_earth_fault_pe_role_without_duplicate_checkbox(
    tmp_path, monkeypatch
):
    monkeypatch.setattr(web, "db", Database(tmp_path / "quick-pe-missing.db"))
    captured = capture_template_context(monkeypatch)
    client = TestClient(web.app)

    response = client.post("/quick", data={
        "input_basis": "current", "input_value": "50", "phase": "3", "voltage_v": "380",
        "conductor_family": "YJV",
        "conductor_configuration": "yjv_4c_3ph_n_pe",
        "installation_scenario": "tray",
        "tray_type": "horizontal_perforated",
        "tray_layers": "1",
        "tray_cables_per_layer": "1",
        "length_m": "50",
        "earth_fault_enabled": "true",
        "earthing_system": "TN-S",
        "nominal_line_to_earth_voltage_v": "230",
        "circuit_application": "distribution",
        "protection_type": "overcurrent",
        "protective_device_characteristic": "mcb_c",
        "fault_fourth_conductor_role": "PE",
        "pe_thermal_enabled": "true",
        "fault_clearing_time_s": "0.1",
    })

    pe = captured["result"]["pe_thermal"]
    assert response.status_code == 200
    assert pe["provisional_status"] == "通过"
    assert pe["outputs"]["protective_conductor_section_mm2"] == 6
    assert "PE导体热稳定" not in captured["result"]["outputs"]["incomplete_checks"]


def test_quick_four_core_does_not_infer_pe_when_role_is_missing(
    tmp_path, monkeypatch
):
    monkeypatch.setattr(web, "db", Database(tmp_path / "quick-four-core-no-pe.db"))
    captured = capture_template_context(monkeypatch)
    client = TestClient(web.app)

    response = client.post("/quick", data={
        "input_basis": "current", "input_value": "50", "phase": "3", "voltage_v": "380",
        "conductor_family": "YJV",
        "conductor_configuration": "yjv_4c_3ph_n_pe",
        "installation_scenario": "tray",
        "tray_type": "horizontal_perforated",
        "tray_layers": "1",
        "tray_cables_per_layer": "1",
        "length_m": "50",
        "earth_fault_enabled": "true",
        "earthing_system": "TN-S",
        "circuit_application": "distribution",
    })

    assert response.status_code == 200
    assert "pe_thermal" not in captured["result"]
    assert captured["form"]["fault_fourth_conductor_role"] == ""
    assert "PE导体热稳定" in captured["result"]["outputs"]["incomplete_checks"]


def test_quick_auto_builds_short_circuit_icu_and_thermal_constraints_without_expert_values(
    tmp_path, monkeypatch
):
    monkeypatch.setattr(web, "db", Database(tmp_path / "quick-auto-protection.db"))
    captured = capture_template_context(monkeypatch)
    client = TestClient(web.app)

    response = client.post("/quick", data={
        "input_basis": "current", "input_value": "50", "phase": "3", "voltage_v": "380",
        "conductor_family": "YJV",
        "conductor_configuration": "yjv_5c_3ph_n_pe",
        "installation_scenario": "tray",
        "tray_type": "horizontal_perforated",
        "tray_layers": "1",
        "tray_cables_per_layer": "1",
        "length_m": "100",
        "transformer_series_code": "scb11",
        "transformer_capacity_kva": "630",
        "transformer_uk_percent": "6",
        "earth_fault_enabled": "true",
        "earthing_system": "TN-S",
        "circuit_application": "distribution",
    })

    result = captured["result"]
    line = result["line_end_short_circuit"]["outputs"]
    phase = result["phase_thermal"]
    pe = result["pe_thermal"]
    stages = {item["code"]: item for item in result["outputs"]["workflow_stages"]}
    assert response.status_code == 200
    assert captured["form"]["short_circuit_line_enabled"] == "true"
    assert captured["form"]["source_impedance_mode"] == "short_circuit_capacity"
    assert captured["form"]["source_short_circuit_capacity_mva"] == "100"
    assert captured["form"]["breaker_installation_point"] == "line_start"
    assert line["terminal_short_circuit_current_ka"] > 0
    assert line["line_start_short_circuit_current_ka"] > line["terminal_short_circuit_current_ka"]
    assert line["required_breaking_capacity_ka"] == line["line_start_short_circuit_current_ka"]
    assert line["required_breaking_capacity_point"] == "line_start"
    assert phase["provisional_status"] == "无法判断"
    assert phase["outputs"]["maximum_permitted_clearing_time_s"] > 0
    assert phase["outputs"]["maximum_permitted_let_through_energy_a2s"] > 0
    assert pe["provisional_status"] == "无法判断"
    assert pe["outputs"]["maximum_permitted_clearing_time_s"] > 0
    assert pe["outputs"]["maximum_permitted_let_through_energy_a2s"] > 0
    assert stages["short_circuit"]["state"] == "completed"
    assert stages["phase_thermal"]["state"] == "candidate"
    assert stages["pe_thermal"]["state"] == "candidate"
    assert "相导体切除时间/I²t实物复核" in result["outputs"]["incomplete_checks"]
    assert "PE切除时间/I²t实物复核" in result["outputs"]["incomplete_checks"]


def test_quick_only_uses_cvs_curve_after_explicit_product_series_selection(
    tmp_path, monkeypatch
):
    monkeypatch.setattr(web, "db", Database(tmp_path / "quick-cvs-review.db"))
    captured = capture_template_context(monkeypatch)
    client = TestClient(web.app)
    base_data = {
        "input_basis": "current", "input_value": "50", "phase": "3",
        "voltage_v": "380", "conductor_family": "YJV",
        "conductor_configuration": "yjv_5c_3ph_n_pe",
        "installation_scenario": "tray", "tray_type": "horizontal_perforated",
        "tray_layers": "1", "tray_cables_per_layer": "1", "length_m": "100",
        "transformer_series_code": "scb11", "transformer_capacity_kva": "630",
        "transformer_uk_percent": "6",
    }

    response = client.post("/quick", data=base_data)
    assert response.status_code == 200
    assert "existing_breaker_product_reference" not in captured["result"]

    response = client.post("/quick", data={
        **base_data,
        "existing_breaker_series": "SCHNEIDER.EASYPACT.CVS.2024",
        "existing_breaker_trip_unit_family": "TM-D",
    })
    product = captured["result"]["existing_breaker_product_reference"]
    thermal = captured["result"]["existing_breaker_phase_thermal"]

    assert response.status_code == 200
    assert product["frame_code"] == "CVS100"
    assert product["rated_current_source"].startswith("复用本次通用断路器初选")
    assert product["icu_ka"] >= thermal["prospective_short_circuit_current_ka"]
    assert thermal["provisional_status"] in {"通过", "不通过"}
    assert thermal["status"] == "无法判断"
    assert "断路器产品参考复核" in response.text
    assert "瞬时脱扣值" in response.text
    assert "性能等级由本次所需Icu自动匹配" in response.text


def test_quick_cvs_review_respects_user_confirmed_nameplate_current(
    tmp_path, monkeypatch
):
    monkeypatch.setattr(web, "db", Database(tmp_path / "quick-cvs-nameplate.db"))
    captured = capture_template_context(monkeypatch)
    client = TestClient(web.app)

    response = client.post("/quick", data={
        "input_basis": "current", "input_value": "100", "phase": "3",
        "voltage_v": "380", "conductor_family": "YJV",
        "conductor_configuration": "yjv_5c_3ph_n_pe",
        "installation_scenario": "tray", "tray_type": "horizontal_perforated",
        "tray_layers": "1", "tray_cables_per_layer": "1", "length_m": "100",
        "transformer_series_code": "scb11", "transformer_capacity_kva": "630",
        "transformer_uk_percent": "6",
        "existing_breaker_series": "SCHNEIDER.EASYPACT.CVS.2024",
        "existing_breaker_rated_current_a": "125",
        "existing_breaker_trip_unit_family": "TM-D",
    })
    product = captured["result"]["existing_breaker_product_reference"]

    assert response.status_code == 200
    assert product["frame_code"] == "CVS160"
    assert product["rated_current_a"] == 125
    assert product["rated_current_source"] == "用户填写的现场设备额定电流"


def test_quick_phase_thermal_uses_line_start_maximum_short_circuit_current(
    tmp_path, monkeypatch
):
    monkeypatch.setattr(web, "db", Database(tmp_path / "quick-phase-thermal.db"))
    captured = capture_template_context(monkeypatch)
    client = TestClient(web.app)

    response = client.post("/quick", data={
        "input_basis": "current", "input_value": "100", "phase": "3", "voltage_v": "400",
        "conductor_family": "YJV", "installation_scenario": "conduit", "length_m": "100",
        "short_circuit_line_enabled": "on",
        "short_circuit_line_type": "cable", "short_circuit_line_section_mm2": "50",
        "transformer_capacity_kva": "1000", "transformer_uk_percent": "6",
        "transformer_pk_kw": "10",
        "source_impedance_mode": "infinite_capacity",
        "voltage_factor_c": "1.05",
        "breaker_installation_point": "line_end", "breaker_icu_ka": "25",
        "phase_thermal_enabled": "true", "fault_clearing_time_s": "0.1",
    })

    phase = captured["result"]["phase_thermal"]
    assert response.status_code == 200
    assert phase["outputs"]["phase_conductor_section_mm2"] > 0
    assert phase["outputs"]["prospective_phase_short_circuit_current_a"] > 0
    assert phase["outputs"]["actual_thermal_stress_a2s"] > 0
    assert phase["provisional_status"] in {"通过", "不通过"}
    assert phase["status"] == "无法判断"
    assert "相导体短路热稳定暂算" in response.text
    assert "相导体热稳定" not in captured["result"]["outputs"]["incomplete_checks"]


def test_quick_phase_thermal_does_not_use_breaker_rated_current_as_fault_current(
    tmp_path, monkeypatch
):
    monkeypatch.setattr(web, "db", Database(tmp_path / "quick-phase-thermal-missing.db"))
    captured = capture_template_context(monkeypatch)
    client = TestClient(web.app)

    response = client.post("/quick", data={
        "input_basis": "current", "input_value": "50", "phase": "3", "voltage_v": "380",
        "conductor_family": "YJV", "installation_scenario": "tray",
        "phase_thermal_enabled": "true", "fault_clearing_time_s": "0.1",
    })

    phase = captured["result"]["phase_thermal"]
    assert response.status_code == 200
    assert phase["provisional_status"] == "无法判断"
    assert any("prospective_fault_current_a" in warning for warning in phase["warnings"])
    assert "相导体热稳定" in captured["result"]["outputs"]["incomplete_checks"]


def test_quick_group_load_and_voltage_drop_flow(tmp_path, monkeypatch):
    monkeypatch.setattr(web, "db", Database(tmp_path / "quick-group.db"))
    client = TestClient(web.app)

    result = client.post("/quick", data={
        "circuit_role": "group_load",
        "input_basis": "kw",
        "input_value": "30",
        "power_definition": "design",
        "phase": "3",
        "voltage_v": "380",
        "load_type_code": "fan_coil",
        "power_factor": "0.8",
        "conductor_family": "YJV",
        "conductor_configuration": "yjv_3c_3ph_pe",
        "installation_scenario": "tray",
        "tray_type": "horizontal_perforated",
        "tray_layers": "1",
        "tray_cables_per_layer": "1",
        "length_m": "100",
        "voltage_drop_limit_pct": "5",
    })

    assert result.status_code == 200
    assert "三相汇总说明" in result.text
    assert "电压降率" in result.text
    assert "推荐最小截面" in result.text
    assert "等待必要参数" in result.text
    assert "所选设备子类不适用于当前相制" not in result.text
