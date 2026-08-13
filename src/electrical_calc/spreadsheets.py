from __future__ import annotations

import json
from io import BytesIO
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from .reports import (
    AUDIT_CHECK_LABELS,
    NETWORK_SEGMENT_LABELS,
    network_derived_rows,
    network_input_rows,
)


HEADERS = [
    ("回路编号", "code", True, "文本，项目内唯一"),
    ("回路名称", "name", True, "文本"),
    ("相制", "phase", True, "1 或 3"),
    ("电压(V)", "voltage_v", True, "内部单位：V"),
    ("安装功率(kW)", "installed_power_kw", True, "内部单位：kW"),
    ("需用系数", "demand_factor", True, "0 < 值 ≤ 1"),
    ("功率因数", "power_factor", True, "0 < 值 ≤ 1"),
    ("效率", "efficiency", True, "0 < 值 ≤ 1"),
    ("线路长度(m)", "length_m", False, "内部单位：m"),
    ("电缆规格", "cable_spec", False, "仅作记录"),
    ("电缆载流量(A)", "cable_ampacity_a", False, "必须来自已核实产品数据"),
    ("线路电阻(Ω/km)", "cable_r_ohm_per_km", False, "必须注明采用工况"),
    ("线路电抗(Ω/km)", "cable_x_ohm_per_km", False, "必须注明采用工况"),
    ("电压降限值(%)", "voltage_drop_limit_pct", False, "须有已批准依据"),
    ("保护器件型号", "breaker_model", False, "仅作记录"),
    ("保护器件额定电流(A)", "breaker_rating_a", False, "来自产品参数"),
    ("分断能力(kA)", "breaking_capacity_ka", False, "来自产品参数"),
    ("电源电阻(Ω)", "source_r_ohm", False, "折算至回路电压侧"),
    ("电源电抗(Ω)", "source_x_ohm", False, "折算至回路电压侧"),
    ("变压器电阻(Ω)", "transformer_r_ohm", False, "折算至回路电压侧"),
    ("变压器电抗(Ω)", "transformer_x_ohm", False, "折算至回路电压侧"),
]

HEADER_MAP = {label: key for label, key, _, _ in HEADERS}
TEXT_FIELDS = {"code", "name", "phase", "cable_spec", "breaker_model"}
REQUIRED_FIELDS = {key for _, key, required, _ in HEADERS if required}

NAVY = "17324D"
TEAL = "008B8B"
PALE = "EAF3F5"
ORANGE = "F4A261"
WHITE = "FFFFFF"
GRID = Side(style="thin", color="D8E1E8")


def _style_title(sheet, last_col: int, title: str, subtitle: str) -> None:
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col)
    sheet["A1"] = title
    sheet["A1"].font = Font(name="Microsoft YaHei", size=18, bold=True, color=WHITE)
    sheet["A1"].fill = PatternFill("solid", fgColor=NAVY)
    sheet["A1"].alignment = Alignment(vertical="center")
    sheet.row_dimensions[1].height = 34
    sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=last_col)
    sheet["A2"] = subtitle
    sheet["A2"].font = Font(name="Microsoft YaHei", size=10, color="52606D")
    sheet["A2"].alignment = Alignment(wrap_text=True, vertical="center")
    sheet.row_dimensions[2].height = 32


def _fit_columns(sheet, widths: list[float]) -> None:
    for index, width in enumerate(widths, 1):
        sheet.column_dimensions[get_column_letter(index)].width = min(width, 30)


def create_input_template() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "回路导入"
    _style_title(
        ws,
        len(HEADERS),
        "电气回路批量导入模板",
        "黄色表头为必填。请勿修改表头；示例数据仅用于说明字段格式，不是设计参数或规范取值。",
    )
    labels = [item[0] for item in HEADERS]
    ws.append([])
    ws.append(labels)
    example = [
        "AL-01", "示例照明回路", "1", 220, 5.0, 0.8, 0.9, 1.0, 35,
        "示例-请替换", None, None, None, None, "示例-请替换", None, None,
        None, None, None, None,
    ]
    ws.append(example)
    for col, (_, _, required, _) in enumerate(HEADERS, 1):
        cell = ws.cell(4, col)
        cell.font = Font(name="Microsoft YaHei", size=10, bold=True, color=WHITE if not required else NAVY)
        cell.fill = PatternFill("solid", fgColor=ORANGE if required else TEAL)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(bottom=GRID)
        ws.cell(5, col).fill = PatternFill("solid", fgColor="FFF8E8")
        ws.cell(5, col).font = Font(name="Microsoft YaHei", italic=True, color="7A5A00")
    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A4:{get_column_letter(len(HEADERS))}5"
    ws.sheet_view.showGridLines = False
    phase_validation = DataValidation(type="list", formula1='"1,3"', allow_blank=False)
    ws.add_data_validation(phase_validation)
    phase_col = labels.index("相制") + 1
    phase_validation.add(f"{get_column_letter(phase_col)}5:{get_column_letter(phase_col)}1000")
    _fit_columns(ws, [14, 20, 8, 12, 16, 12, 12, 10, 14, 18] + [17] * 11)

    guide = wb.create_sheet("字段说明")
    _style_title(guide, 4, "字段说明", "内部单位固定；缺失、不适用与数值 0 必须区分。")
    guide.append([])
    guide.append(["字段", "必填", "内部字段", "规则"])
    for label, key, required, description in HEADERS:
        guide.append([label, "是" if required else "否", key, description])
    for cell in guide[4]:
        cell.fill = PatternFill("solid", fgColor=NAVY)
        cell.font = Font(name="Microsoft YaHei", bold=True, color=WHITE)
    guide.freeze_panes = "A5"
    guide.sheet_view.showGridLines = False
    _fit_columns(guide, [24, 10, 28, 52])
    for row in guide.iter_rows(min_row=5):
        row[3].alignment = Alignment(wrap_text=True, vertical="top")

    stream = BytesIO()
    wb.save(stream)
    return stream.getvalue()


def parse_circuit_workbook(content: bytes) -> tuple[list[dict[str, Any]], list[str]]:
    errors: list[str] = []
    rows: list[dict[str, Any]] = []
    try:
        wb = load_workbook(BytesIO(content), data_only=True, read_only=True)
    except Exception as exc:
        return [], [f"无法读取 Excel 文件：{exc}"]
    ws = wb["回路导入"] if "回路导入" in wb.sheetnames else wb.active
    header_row = None
    mapping: dict[int, str] = {}
    for row_index in range(1, min(ws.max_row, 15) + 1):
        candidates = {index: HEADER_MAP.get(str(cell.value).strip()) for index, cell in enumerate(ws[row_index], 1) if cell.value is not None}
        if "code" in candidates.values() and "name" in candidates.values():
            header_row = row_index
            mapping = {index: key for index, key in candidates.items() if key}
            break
    if header_row is None:
        return [], ["未找到有效表头；请使用平台下载的导入模板。"]
    missing_headers = REQUIRED_FIELDS - set(mapping.values())
    if missing_headers:
        return [], ["缺少必填列：" + "、".join(sorted(missing_headers))]

    seen: set[str] = set()
    for row_index in range(header_row + 1, ws.max_row + 1):
        values = {key: ws.cell(row_index, col).value for col, key in mapping.items()}
        if not any(value not in (None, "") for value in values.values()):
            continue
        normalized: dict[str, Any] = {}
        row_errors: list[str] = []
        for key in HEADER_MAP.values():
            value = values.get(key)
            if key in TEXT_FIELDS:
                normalized[key] = "" if value is None else str(value).strip()
            elif value in (None, ""):
                normalized[key] = None
            else:
                try:
                    normalized[key] = float(value)
                except (TypeError, ValueError):
                    row_errors.append(f"{key} 不是有效数值")
        code = normalized.get("code", "")
        if not code:
            row_errors.append("回路编号不能为空")
        elif code in seen:
            row_errors.append(f"回路编号 {code} 在文件内重复")
        seen.add(code)
        for field in REQUIRED_FIELDS:
            if normalized.get(field) in (None, ""):
                row_errors.append(f"{field} 为必填项")
        if normalized.get("phase") not in {"1", "3"}:
            row_errors.append("相制必须为 1 或 3")
        for field in ("demand_factor", "power_factor", "efficiency"):
            value = normalized.get(field)
            if value is not None and not 0 < value <= 1:
                row_errors.append(f"{field} 必须大于 0 且不大于 1")
        nonnegative = [
            "voltage_v", "installed_power_kw", "length_m", "cable_ampacity_a",
            "cable_r_ohm_per_km", "cable_x_ohm_per_km", "voltage_drop_limit_pct",
            "breaker_rating_a", "breaking_capacity_ka", "source_r_ohm", "source_x_ohm",
            "transformer_r_ohm", "transformer_x_ohm",
        ]
        for field in nonnegative:
            value = normalized.get(field)
            if value is not None and value < 0:
                row_errors.append(f"{field} 不能为负数")
        if row_errors:
            errors.extend(f"第 {row_index} 行：{message}" for message in row_errors)
        else:
            rows.append(normalized)
    if not rows and not errors:
        errors.append("文件中没有可导入的数据行。")
    return rows, errors


def create_project_export(
    project: dict[str, Any],
    circuits: list[dict[str, Any]],
    runs: list[dict[str, Any]],
    rules: list[dict[str, Any]],
) -> bytes:
    wb = Workbook()
    summary = wb.active
    summary.title = "项目汇总"
    _style_title(summary, 4, f"{project['name']} - 电气计算成果", "本文件为计算快照；正式性取决于各计算依据的批准状态。")
    summary.append([])
    summary.append(["项目编号", project["code"], "项目名称", project["name"]])
    summary.append(["回路数量", len(circuits), "计算记录", len(runs)])
    summary.append(["导出说明", "未批准依据的结果均为暂算，不能作为正式设计结论。", "", ""])
    summary.merge_cells("B7:D7")
    for row in range(4, 8):
        summary.cell(row, 1).font = Font(bold=True, color=NAVY)
    summary.sheet_view.showGridLines = False
    _fit_columns(summary, [18, 32, 18, 32])

    circuits_ws = wb.create_sheet("回路输入快照")
    circuit_headers = [label for label, _, _, _ in HEADERS] + ["数据修订版"]
    circuits_ws.append(circuit_headers)
    for circuit in circuits:
        circuits_ws.append([circuit.get(key) for _, key, _, _ in HEADERS] + [circuit.get("revision")])
    _format_table(circuits_ws, len(circuit_headers))

    result_ws = wb.create_sheet("计算记录")
    result_headers = ["记录ID", "回路编号", "模块", "正式状态", "暂算状态", "是否过期", "引擎版本", "计算时间"]
    result_ws.append(result_headers)
    for run in runs:
        result_ws.append(
            [
                run["id"], run["circuit_code"], run["module"], run["status"],
                run["provisional_status"], "是" if run["stale"] else "否",
                run["engine_version"], run["created_at"],
            ]
        )
    _format_table(result_ws, len(result_headers))

    detail_ws = wb.create_sheet("结果明细")
    detail_headers = ["记录ID", "回路编号", "模块", "结果项", "结果值", "正式状态", "暂算状态", "是否过期"]
    detail_ws.append(detail_headers)
    process_ws = wb.create_sheet("计算过程")
    process_headers = ["记录ID", "回路编号", "模块", "步骤序号", "步骤", "表达式", "结果", "单位"]
    process_ws.append(process_headers)
    warning_ws = wb.create_sheet("警告清单")
    warning_headers = ["记录ID", "回路编号", "模块", "正式状态", "是否过期", "警告"]
    warning_ws.append(warning_headers)
    snapshot_ws = wb.create_sheet("计算输入快照")
    snapshot_headers = ["记录ID", "回路编号", "模块", "字段", "值", "回路修订"]
    snapshot_ws.append(snapshot_headers)
    for run in runs:
        result_data = json.loads(run["result_json"]) if isinstance(run.get("result_json"), str) else run.get("result_json", {})
        process_data = json.loads(run["process_json"]) if isinstance(run.get("process_json"), str) else run.get("process_json", [])
        warning_data = json.loads(run["warnings_json"]) if isinstance(run.get("warnings_json"), str) else run.get("warnings_json", [])
        input_data = json.loads(run["input_snapshot"]) if isinstance(run.get("input_snapshot"), str) else run.get("input_snapshot", {})
        for key, value in result_data.items():
            detail_ws.append([run["id"], run["circuit_code"], run["module"], key, value, run["status"], run["provisional_status"], "是" if run["stale"] else "否"])
        for index, step in enumerate(process_data, 1):
            process_ws.append([run["id"], run["circuit_code"], run["module"], index, step.get("label"), step.get("expression"), step.get("value"), step.get("unit")])
        for warning in warning_data:
            warning_ws.append([run["id"], run["circuit_code"], run["module"], run["status"], "是" if run["stale"] else "否", warning])
        for key, value in input_data.items():
            if key not in {"id", "project_id", "created_at", "updated_at"}:
                snapshot_ws.append([run["id"], run["circuit_code"], run["module"], key, value, run.get("circuit_revision")])
    _format_table(detail_ws, len(detail_headers))
    _format_table(process_ws, len(process_headers))
    _format_table(warning_ws, len(warning_headers))
    _format_table(snapshot_ws, len(snapshot_headers))
    warning_ws.column_dimensions["F"].width = 60
    process_ws.column_dimensions["F"].width = 48

    rule_ws = wb.create_sheet("依据清单")
    rule_headers = ["依据编号", "名称", "状态", "文件", "版本", "条文号", "原文", "页码", "备注"]
    rule_ws.append(rule_headers)
    for rule in rules:
        rule_ws.append(
            [
                rule["code"], rule["name"], rule["status"], rule["document_name"],
                rule["document_version"], rule["clause_no"], rule["original_text"],
                rule["page_no"], rule["note"],
            ]
        )
    _format_table(rule_ws, len(rule_headers))
    rule_ws.column_dimensions["G"].width = 60
    for row in rule_ws.iter_rows(min_row=2):
        row[6].alignment = Alignment(wrap_text=True, vertical="top")

    stream = BytesIO()
    wb.save(stream)
    return stream.getvalue()


def create_network_run_export(run: dict[str, Any]) -> bytes:
    """从完整回路版本快照生成成果表，不在导出阶段重新计算。"""

    wb = Workbook()
    summary = wb.active
    summary.title = "成果总览"
    input_data = run["input_snapshot"]
    result = run["result_json"]
    outputs = result.get("outputs", {})
    choices = outputs.get("viable_combinations") or outputs.get("incomplete_combinations") or []
    choice = choices[0] if choices else {}
    chain = choice.get("chain_result", {}).get("outputs", {})
    _style_title(
        summary,
        4,
        f"{run['project_name']} - 完整低压回路成果",
        "本文件来自不可覆盖的计算版本快照；未批准依据的暂算结果不能作为正式设计结论。",
    )
    summary.append([])
    summary.append(["项目编号", run["project_code"], "项目名称", run["project_name"]])
    summary.append(["回路编号", input_data.get("circuit_code"), "回路名称", input_data.get("circuit_name")])
    summary.append(["计算版本", f"V{run['network_revision']} / #{run['id']}", "任务", "既有核验" if run["task_mode"] == "audit" else "快速设计"])
    summary.append(["正式状态", run["status"], "暂算状态", run["provisional_status"]])
    summary.append(["是否过期", "是" if run["stale"] else "否", "引擎版本", run["engine_version"]])
    summary.append(["计算时间", run["created_at"], "输入快照", "已冻结"])
    for row in range(4, 10):
        summary.cell(row, 1).font = Font(bold=True, color=NAVY)
        summary.cell(row, 3).font = Font(bold=True, color=NAVY)
    summary.sheet_view.showGridLines = False
    _fit_columns(summary, [18, 34, 18, 34])

    inputs = wb.create_sheet("输入与推导")
    inputs.append(["类别", "字段", "值"])
    for label, value in network_input_rows(input_data):
        inputs.append(["用户输入", label, value])
    for label, value in network_derived_rows(run["derived_json"]):
        display_value = (
            json.dumps(value, ensure_ascii=False)
            if isinstance(value, (dict, list, tuple))
            else value
        )
        inputs.append(["系统推导", label, display_value])
    _format_table(inputs, 3)
    _fit_columns(inputs, [16, 40, 42])

    cables = wb.create_sheet("电缆方案")
    cables.append(["线路段", "电缆规格", "计算电流Ib(A)", "修正载流量Iz(A)", "状态"])
    for cable in choice.get("cables", []):
        cables.append([
            NETWORK_SEGMENT_LABELS.get(
                str(cable.get("candidate_id", "")).split(":")[0],
                str(cable.get("candidate_id", "")).split(":")[0],
            ),
            cable.get("cable_specification"),
            cable.get("minimum_required_ampacity_a"),
            cable.get("corrected_ampacity_a"),
            cable.get("provisional_status", ""),
        ])
    _format_table(cables, 5)
    _fit_columns(cables, [22, 42, 20, 22, 16])

    breakers = wb.create_sheet("保护器件")
    breakers.append(["序号", "类别", "额定电流In(A)", "壳架电流(A)", "额定电压Ue(V)", "分断能力Icu(kA)", "状态"])
    for index, breaker in enumerate(choice.get("breakers", []), 1):
        breakers.append([
            index, breaker.get("family"), breaker.get("rated_current_a"),
            breaker.get("frame_current_a"), breaker.get("rated_voltage_v"),
            breaker.get("selected_icu_ka"), breaker.get("provisional_status"),
        ])
    _format_table(breakers, 7)
    _fit_columns(breakers, [10, 18, 20, 20, 20, 22, 16])

    audit_outputs = run.get("audit_json", {}).get("outputs", {})
    if audit_outputs:
        audit = wb.create_sheet("原设计核验")
        audit.append(["对象", "原规格/标识", "核验项", "判定", "判定条件", "缺失/说明"])
        for component in audit_outputs.get("component_matrix", []):
            for code, check in component.get("checks", {}).items():
                audit.append([
                    component.get("component_name", component.get("component_type")),
                    component.get("designation"), AUDIT_CHECK_LABELS.get(code, code),
                    check.get("status", "无法判断"), check.get("criterion", ""),
                    check.get("reason", "") or "；".join(component.get("remediation_actions", [])),
                ])
        for check in audit_outputs.get("cross_component_checks", []):
            audit.append([
                "跨部件配合", check.get("check_name"), "系统配合",
                check.get("status", "无法判断"), check.get("criterion", ""), check.get("reason", ""),
            ])
        _format_table(audit, 6)
        _fit_columns(audit, [18, 46, 24, 16, 58, 58])

    nodes = wb.create_sheet("逐节点校核")
    nodes.append(["节点", "累计压降(%)", "最大三相短路(kA)", "最小相-PE故障(A)"])
    for node in chain.get("node_results", []):
        nodes.append([
            node.get("node_name"), node.get("cumulative_voltage_drop_percent"),
            node.get("three_phase_short_circuit_ka"), node.get("earth_fault_current_a"),
        ])
    _format_table(nodes, 4)
    _fit_columns(nodes, [34, 22, 24, 25])

    warnings = wb.create_sheet("警告与未闭合项")
    warnings.append(["类型", "内容"])
    for warning in run.get("warnings_json", []):
        warnings.append(["计算警告", warning])
    for item in choice.get("missing_items", []):
        warnings.append(["未闭合项", item])
    _format_table(warnings, 2)
    _fit_columns(warnings, [18, 90])
    for row in warnings.iter_rows(min_row=2):
        row[1].alignment = Alignment(wrap_text=True, vertical="top")

    rules = wb.create_sheet("依据快照")
    rules.append(["依据编号", "名称", "状态", "文件", "条文/表号", "页码", "原文"])
    for code, rule in run["rule_snapshot"].items():
        rules.append([
            code, rule.get("name"), rule.get("status"), rule.get("document_name"),
            rule.get("clause_no"), rule.get("page_no"), rule.get("original_text"),
        ])
    _format_table(rules, 7)
    _fit_columns(rules, [34, 28, 14, 38, 24, 24, 90])
    for row in rules.iter_rows(min_row=2):
        row[6].alignment = Alignment(wrap_text=True, vertical="top")

    stream = BytesIO()
    wb.save(stream)
    return stream.getvalue()


def create_motor_run_export(run: dict[str, Any]) -> bytes:
    """从不可变电动机快照生成Excel成果。"""
    wb = Workbook()
    summary = wb.active; summary.title = "成果总览"
    result = run["result_json"]; load = result.get("load", {}).get("outputs", {})
    candidate = result.get("recommended_candidate") or {}; scheme = candidate.get("primary_scheme") or {}
    summary.append(["项目", run["project_name"], "回路", run["input_snapshot"].get("circuit_code")])
    summary.append(["计算版本", f"V{run['motor_revision']} / #{run['id']}", "引擎", run["engine_version"]])
    summary.append(["正式状态", run["status"], "暂算状态", run["provisional_status"]])
    summary.append(["额定电流(A)", load.get("rated_current_a"), "启动电流(A)", load.get("starting_current_a")])
    summary.append(["主电缆", candidate.get("cable", {}).get("cable_specification"), "保护器", scheme.get("breaker")])
    summary.append(["接触器", scheme.get("contactor"), "过载保护", scheme.get("overload_device")])
    _fit_columns(summary, [20, 52, 20, 52])
    inputs=wb.create_sheet("输入快照"); inputs.append(["字段","值"])
    for key,value in run["input_snapshot"].items(): inputs.append([key,value])
    _format_table(inputs,2); _fit_columns(inputs,[38,60])
    checks=wb.create_sheet("校核结果"); checks.append(["校核","结果"])
    for item in scheme.get("closed_checks",[]): checks.append([item,"已闭合"])
    for item in scheme.get("professional_pending",[]): checks.append([item,"待深化"])
    _format_table(checks,2); _fit_columns(checks,[55,20])
    rules=wb.create_sheet("依据快照"); rules.append(["编号","状态","文件","条文/表号","页码"])
    for code,rule in run["rule_snapshot"].items(): rules.append([code,rule.get("status"),rule.get("document_name"),rule.get("clause_no"),rule.get("page_no")])
    _format_table(rules,5); _fit_columns(rules,[38,16,45,30,28])
    stream=BytesIO(); wb.save(stream); return stream.getvalue()


def create_drawing_project_export(
    project: dict[str, Any], circuits: list[dict[str, Any]], summary: dict[str, Any],
    settings: dict[str, Any],
) -> bytes:
    """导出项目内所有图纸回路的当前有效核验结果。"""
    wb = Workbook()
    overview = wb.active; overview.title = "项目汇总"
    overview.append(["项目编号", project.get("code"), "项目名称", project.get("name")])
    overview.append(["有效回路数", summary.get("circuit_count"), "Ib算术合计(A)", summary.get("arithmetic_total_current_a")])
    overview.append(["同时系数", summary.get("simultaneity_factor"), "上游计算电流(A)", summary.get("upstream_design_current_a")])
    overview.append(["变压器额定电流(A)", summary.get("transformer_rated_current_a"), "容量复核", summary.get("transformer_capacity_status")])
    overview.append(["系数来源/说明", settings.get("source_note", ""), "电源树一致", "是" if summary.get("source_consistent") else "否"])
    completeness = summary.get("completeness", {})
    overview.append(["工程数据闭合", completeness.get("engineering_data_gate"), "正式成果发布", completeness.get("formal_release_gate")])
    overview.append(["不通过项", completeness.get("counts", {}).get("不通过", 0), "无法判断项", completeness.get("counts", {}).get("无法判断", 0)])
    _fit_columns(overview, [24, 42, 24, 42])

    sheet = wb.create_sheet("回路清单")
    sheet.append(["回路编号", "回路名称", "修订", "计算电流(A)", "正式状态", "暂算状态", "计算时间"])
    for item in circuits:
        sheet.append([item.get("circuit_code"), item.get("circuit_name"), item.get("revision"),
                      item.get("derived_json", {}).get("design_current_a"), item.get("status") or "未计算",
                      item.get("provisional_status") or "未计算", item.get("calculated_at") or ""])
    _format_table(sheet, 7); _fit_columns(sheet, [22, 38, 12, 20, 18, 18, 24])

    groups = wb.create_sheet("上游配电分组")
    groups.append(["层级", "变压器编号", "母线段编号", "馈线柜编号", "直接下级合计(A)", "系数", "本级设计电流(A)", "设备额定电流(A)", "负荷复核", "Ikmax(kA)", "Icw(1s)(kA)", "Icw复核", "断路器", "Icu(kA)", "Icu复核", "选择性极限(kA)", "选择性复核", "来源说明"])
    for level, key in (("馈线柜", "feeder_cabinet_groups"), ("母线段", "bus_section_groups"), ("变压器", "transformer_groups")):
        for group in summary.get(key, []):
            codes = list(group["codes"]) + [""] * (3-len(group["codes"]))
            direct = group.get("arithmetic_total_current_a") if level == "馈线柜" else group.get("direct_child_current_a")
            groups.append([level, *codes, direct, group.get("factor"), group.get("design_current_a"),
                           group.get("rated_current_a"), group.get("equipment_status"),
                           group.get("prospective_short_circuit_ka"), group.get("short_time_withstand_ka"),
                           group.get("short_time_withstand_status"), group.get("breaker_designation"),
                           group.get("breaker_breaking_capacity_ka"), group.get("breaking_capacity_status"),
                           group.get("selectivity_limit_ka"), group.get("selectivity_status"),
                           "；".join(filter(None, (group.get("source_note", ""), group.get("selectivity_reference", ""))))])
    _format_table(groups, 18); _fit_columns(groups, [14, 18, 18, 18, 20, 12, 20, 20, 16, 18, 18, 16, 28, 16, 16, 20, 16, 55])

    issues = wb.create_sheet("问题清单")
    issues.append(["优先级", "范围", "对象", "校核项", "状态", "处理要求", "说明/条件"])
    for issue in completeness.get("issues", []):
        issues.append([issue.get("priority"), issue.get("scope"), issue.get("subject"),
                       issue.get("check"), issue.get("status"), issue.get("action"), issue.get("detail")])
    _format_table(issues, 7); _fit_columns(issues, [14, 16, 50, 28, 16, 75, 55])
    for row in issues.iter_rows(min_row=2):
        for cell in row: cell.alignment = Alignment(wrap_text=True, vertical="top")

    checks = wb.create_sheet("逐回路部件核验")
    checks.append(["回路", "部件", "规格/标识", "核验项", "判定", "判定条件", "原因/整改"])
    for item in circuits:
        outputs = item.get("audit_json", {}).get("outputs", {})
        for component in outputs.get("component_matrix", []):
            actions = "；".join(component.get("remediation_actions", []))
            for code, check in component.get("checks", {}).items():
                checks.append([item.get("circuit_code"), component.get("component_name", component.get("component_type")),
                               component.get("designation"), AUDIT_CHECK_LABELS.get(code, code), check.get("status"),
                               check.get("criterion", ""), check.get("reason", "") or actions])
        for check in outputs.get("cross_component_checks", []):
            checks.append([item.get("circuit_code"), "跨部件配合", check.get("check_name"), "系统配合",
                           check.get("status"), check.get("criterion", ""), check.get("reason", "")])
    _format_table(checks, 7); _fit_columns(checks, [18, 25, 42, 22, 16, 55, 65])
    for row in checks.iter_rows(min_row=2):
        for cell in row: cell.alignment = Alignment(wrap_text=True, vertical="top")

    warnings = wb.create_sheet("警告")
    warnings.append(["范围", "内容"])
    for warning in summary.get("warnings", []): warnings.append(["项目汇总", warning])
    for item in circuits:
        for warning in item.get("audit_json", {}).get("warnings", []): warnings.append([item.get("circuit_code"), warning])
    _format_table(warnings, 2); _fit_columns(warnings, [20, 100])
    stream = BytesIO(); wb.save(stream); return stream.getvalue()


def _format_table(ws, last_col: int) -> None:
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(last_col)}{max(ws.max_row, 1)}"
    ws.sheet_view.showGridLines = False
    for cell in ws[1]:
        cell.fill = PatternFill("solid", fgColor=NAVY)
        cell.font = Font(name="Microsoft YaHei", bold=True, color=WHITE)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row_index in range(2, ws.max_row + 1):
        if row_index % 2 == 0:
            for cell in ws[row_index]:
                cell.fill = PatternFill("solid", fgColor=PALE)
    for column in range(1, last_col + 1):
        values = [str(ws.cell(row, column).value or "") for row in range(1, min(ws.max_row, 60) + 1)]
        ws.column_dimensions[get_column_letter(column)].width = min(max(len(value) for value in values) * 1.2 + 2, 24)
